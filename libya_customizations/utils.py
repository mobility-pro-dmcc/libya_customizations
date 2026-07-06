import json
import frappe
from frappe import _
from erpnext.controllers.accounts_controller import validate_and_delete_children, set_order_defaults
from frappe.model.workflow import get_workflow_name, is_transition_condition_satisfied
from frappe.utils import (flt, get_link_to_form, getdate)

from erpnext.buying.utils import update_last_purchase_rate
from erpnext.stock.doctype.packed_item.packed_item import make_packing_list
from erpnext.stock.get_item_details import get_conversion_factor
from erpnext.accounts.doctype.payment_entry.payment_entry import get_outstanding_reference_documents
from erpnext.accounts.doctype.unreconcile_payment.unreconcile_payment import get_linked_payments_for_doc
from erpnext.accounts.doctype.unreconcile_payment.unreconcile_payment import create_unreconcile_doc_for_selection

@frappe.whitelist()
def get_linked_document(linked_doctype, docname, linked_field, field):
	return frappe.db.get_value(linked_doctype, [[linked_field, "=", docname]], field)

@frappe.whitelist()
def update_child_qty_rate(parent_doctype, trans_items, parent_doctype_name, child_docname="items"):
	def check_doc_permissions(doc, perm_type="create"):
		try:
			doc.check_permission(perm_type)
		except frappe.PermissionError:
			actions = {"create": "add", "write": "update"}

			frappe.throw(
				_("You do not have permissions to {} items in a {}.").format(
					actions[perm_type], parent_doctype
				),
				title=_("Insufficient Permissions"),
			)

	def validate_workflow_conditions(doc):
		workflow = get_workflow_name(doc.doctype)
		if not workflow:
			return

		workflow_doc = frappe.get_doc("Workflow", workflow)
		current_state = doc.get(workflow_doc.workflow_state_field)
		roles = frappe.get_roles()

		transitions = []
		for transition in workflow_doc.transitions:
			if transition.next_state == current_state and transition.allowed in roles:
				if not is_transition_condition_satisfied(transition, doc):
					continue
				transitions.append(transition.as_dict())

		if not transitions:
			frappe.throw(
				_("You are not allowed to update as per the conditions set in {} Workflow.").format(
					get_link_to_form("Workflow", workflow)
				),
				title=_("Insufficient Permissions"),
			)

	def get_new_child_item(item_row):
		child_doctype = "Sales Order Item" if parent_doctype == "Sales Order" else "Purchase Order Item"
		return set_order_defaults(parent_doctype, parent_doctype_name, child_doctype, child_docname, item_row)

	def validate_quantity(child_item, new_data):
		if not flt(new_data.get("qty")):
			frappe.throw(
				_("Row # {0}: Quantity for Item {1} cannot be zero").format(
					new_data.get("idx"), frappe.bold(new_data.get("item_code"))
				),
				title=_("Invalid Qty"),
			)

		if parent_doctype == "Sales Order" and flt(new_data.get("qty")) < flt(child_item.delivered_qty):
			frappe.throw(_("Cannot set quantity less than delivered quantity"))

		if parent_doctype == "Purchase Order" and flt(new_data.get("qty")) < flt(child_item.received_qty):
			frappe.throw(_("Cannot set quantity less than received quantity"))

	def should_update_supplied_items(doc) -> bool:
		"""Subcontracted PO can allow following changes *after submit*:

		1. Change rate of subcontracting - regardless of other changes.
		2. Change qty and/or add new items and/or remove items
				Exception: Transfer/Consumption is already made, qty change not allowed.
		"""

		supplied_items_processed = any(
			item.supplied_qty or item.consumed_qty or item.returned_qty for item in doc.supplied_items
		)

		update_supplied_items = any_qty_changed or items_added_or_removed or any_conversion_factor_changed
		if update_supplied_items and supplied_items_processed:
			frappe.throw(_("Item qty can not be updated as raw materials are already processed."))

		return update_supplied_items

	def validate_fg_item_for_subcontracting(new_data, is_new):
		if is_new:
			if not new_data.get("fg_item"):
				frappe.throw(
					_("Finished Good Item is not specified for service item {0}").format(
						new_data["item_code"]
					)
				)
			else:
				is_sub_contracted_item, default_bom = frappe.db.get_value(
					"Item", new_data["fg_item"], ["is_sub_contracted_item", "default_bom"]
				)

				if not is_sub_contracted_item:
					frappe.throw(
						_("Finished Good Item {0} must be a sub-contracted item").format(new_data["fg_item"])
					)
				elif not default_bom:
					frappe.throw(_("Default BOM not found for FG Item {0}").format(new_data["fg_item"]))

		if not new_data.get("fg_item_qty"):
			frappe.throw(_("Finished Good Item {0} Qty can not be zero").format(new_data["fg_item"]))

	data = json.loads(trans_items)

	any_qty_changed = False  # updated to true if any item's qty changes
	items_added_or_removed = False  # updated to true if any new item is added or removed
	any_conversion_factor_changed = False

	parent = frappe.get_doc(parent_doctype, parent_doctype_name)

	check_doc_permissions(parent, "write")
	_removed_items = validate_and_delete_children(parent, data)
	items_added_or_removed |= _removed_items

	for d in data:
		new_child_flag = False

		if not d.get("item_code"):
			# ignore empty rows
			continue

		if not d.get("docname"):
			new_child_flag = True
			items_added_or_removed = True
			check_doc_permissions(parent, "create")
			child_item = get_new_child_item(d)
		else:
			check_doc_permissions(parent, "write")
			child_item = frappe.get_doc(parent_doctype + " Item", d.get("docname"))

			prev_rate, new_rate = flt(child_item.get("rate")), flt(d.get("rate"))
			prev_qty, new_qty = flt(child_item.get("qty")), flt(d.get("qty"))
			prev_fg_qty, new_fg_qty = flt(child_item.get("fg_item_qty")), flt(d.get("fg_item_qty"))
			prev_con_fac, new_con_fac = (
				flt(child_item.get("conversion_factor")),
				flt(d.get("conversion_factor")),
			)
			prev_uom, new_uom = child_item.get("uom"), d.get("uom")

			if parent_doctype == "Sales Order":
				prev_date, new_date = child_item.get("delivery_date"), d.get("delivery_date")
			elif parent_doctype == "Purchase Order":
				prev_date, new_date = child_item.get("schedule_date"), d.get("schedule_date")

			rate_unchanged = prev_rate == new_rate
			qty_unchanged = prev_qty == new_qty
			fg_qty_unchanged = prev_fg_qty == new_fg_qty
			uom_unchanged = prev_uom == new_uom
			conversion_factor_unchanged = prev_con_fac == new_con_fac
			any_conversion_factor_changed |= not conversion_factor_unchanged
			date_unchanged = (
				prev_date == getdate(new_date) if prev_date and new_date else False
			)  # in case of delivery note etc
			if (
				rate_unchanged
				and qty_unchanged
				and fg_qty_unchanged
				and conversion_factor_unchanged
				and uom_unchanged
				and date_unchanged
			):
				continue

		validate_quantity(child_item, d)
		if flt(child_item.get("qty")) != flt(d.get("qty")):
			any_qty_changed = True

		if (
			parent.doctype == "Purchase Order"
			and parent.is_subcontracted
			and not parent.is_old_subcontracting_flow
		):
			validate_fg_item_for_subcontracting(d, new_child_flag)
			child_item.fg_item_qty = flt(d["fg_item_qty"])

			if new_child_flag:
				child_item.fg_item = d["fg_item"]

		child_item.qty = flt(d.get("qty"))
		rate_precision = child_item.precision("rate") or 2
		conv_fac_precision = child_item.precision("conversion_factor") or 2
		qty_precision = child_item.precision("qty") or 2

		# Amount cannot be lesser than billed amount, except for negative amounts
		row_rate = flt(d.get("rate"), rate_precision)
		amount_below_billed_amt = flt(child_item.billed_amt, rate_precision) > flt(
			row_rate * flt(d.get("qty"), qty_precision), rate_precision
		)
		if amount_below_billed_amt and row_rate > 0.0:
			frappe.throw(
				_("Row #{0}: Cannot set Rate if amount is greater than billed amount for Item {1}.").format(
					child_item.idx, child_item.item_code
				)
			)
		else:
			child_item.rate = row_rate

		if d.get("conversion_factor"):
			if child_item.stock_uom == child_item.uom:
				child_item.conversion_factor = 1
			else:
				child_item.conversion_factor = flt(d.get("conversion_factor"), conv_fac_precision)

		if d.get("uom"):
			child_item.uom = d.get("uom")
			conversion_factor = flt(
				get_conversion_factor(child_item.item_code, child_item.uom).get("conversion_factor")
			)
			child_item.conversion_factor = (
				flt(d.get("conversion_factor"), conv_fac_precision) or conversion_factor
			)

		if d.get("delivery_date") and parent_doctype == "Sales Order":
			child_item.delivery_date = d.get("delivery_date")

		if d.get("schedule_date") and parent_doctype == "Purchase Order":
			child_item.schedule_date = d.get("schedule_date")

		if flt(child_item.price_list_rate):
			if flt(child_item.rate) > flt(child_item.price_list_rate):
				#  if rate is greater than price_list_rate, set margin
				#  or set discount
				child_item.discount_percentage = 0
				child_item.margin_type = "Amount"
				child_item.margin_rate_or_amount = flt(
					child_item.rate - child_item.price_list_rate,
					child_item.precision("margin_rate_or_amount"),
				)
				child_item.rate_with_margin = child_item.rate
			else:
				child_item.discount_percentage = flt(
					(1 - flt(child_item.rate) / flt(child_item.price_list_rate)) * 100.0,
					child_item.precision("discount_percentage"),
				)
				child_item.discount_amount = flt(child_item.price_list_rate) - flt(child_item.rate)
				child_item.margin_type = ""
				child_item.margin_rate_or_amount = 0
				child_item.rate_with_margin = 0
		if d.get("brand"):
			child_item.brand = d.get("brand")
		if d.get("production_year"):
			child_item.production_year = d.get("production_year")
		child_item.flags.ignore_validate_update_after_submit = True
		if new_child_flag:
			parent.load_from_db()
			child_item.idx = len(parent.items) + 1
			child_item.insert()
		else:
			child_item.save()

	parent.reload()
	parent.flags.ignore_validate_update_after_submit = True
	parent.set_qty_as_per_stock_uom()
	parent.calculate_taxes_and_totals()
	parent.set_total_in_words()
	if parent_doctype == "Sales Order":
		make_packing_list(parent)
		parent.set_gross_profit()
	frappe.get_doc("Authorization Control").validate_approving_authority(
		parent.doctype, parent.company, parent.base_grand_total
	)

	parent.set_payment_schedule()
	if parent_doctype == "Purchase Order":
		parent.validate_minimum_order_qty()
		parent.validate_budget()
		if parent.is_against_so():
			parent.update_status_updater()
	else:
		parent.check_credit_limit()

	# reset index of child table
	for idx, row in enumerate(parent.get(child_docname), start=1):
		row.idx = idx

	parent.save()

	if parent_doctype == "Purchase Order":
		update_last_purchase_rate(parent, is_submit=1)

		if any_qty_changed or items_added_or_removed or any_conversion_factor_changed:
			parent.update_prevdoc_status()

		parent.update_requested_qty()
		parent.update_ordered_qty()
		parent.update_ordered_and_reserved_qty()
		parent.update_receiving_percentage()

		if parent.is_subcontracted:
			if parent.is_old_subcontracting_flow:
				if should_update_supplied_items(parent):
					parent.update_reserved_qty_for_subcontract()
					parent.create_raw_materials_supplied()
				parent.save()
			else:
				if not parent.can_update_items():
					frappe.throw(
						_(
							"Items cannot be updated as Subcontracting Order is created against the Purchase Order {0}."
						).format(frappe.bold(parent.name))
					)
	else:  # Sales Order
		parent.validate_warehouse()
		parent.update_reserved_qty()
		parent.update_project()
		parent.update_prevdoc_status("submit")
		parent.update_delivery_status()

	parent.reload()
	validate_workflow_conditions(parent)

	parent.update_blanket_order()
	parent.update_billing_percentage()
	parent.set_status()

	parent.validate_uom_is_integer("uom", "qty")
	parent.validate_uom_is_integer("stock_uom", "stock_qty")

	# Cancel and Recreate Stock Reservation Entries.
	if parent_doctype == "Sales Order":
		from erpnext.stock.doctype.stock_reservation_entry.stock_reservation_entry import (
			cancel_stock_reservation_entries,
			has_reserved_stock,
		)

		if has_reserved_stock(parent.doctype, parent.name):
			cancel_stock_reservation_entries(parent.doctype, parent.name)

			if parent.per_picked == 0:
				parent.create_stock_reservation_entries()




# excel sheets
from frappe.utils.xlsxutils import INVALID_TITLE_REGEX, ILLEGAL_CHARACTERS_RE, handle_html
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO

def make_xlsx(data, sheet_name, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = Workbook()

	# Sanitize sheet name
	sheet_name_sanitized = INVALID_TITLE_REGEX.sub(" ", sheet_name)
	ws = wb.create_sheet(sheet_name_sanitized, 0)

	# Define green fill for the first row
	green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green

	# Apply green fill to the first row

	# Set bold font for the first row
	row1 = ws.row_dimensions[1]
	row1.font = Font(name="Calibri", bold=True)

	# Process data and write to the sheet
	for row in data:
		clean_row = []
		for item in row:
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = ILLEGAL_CHARACTERS_RE.sub("", value)

			clean_row.append(value)

		ws.append(clean_row)

	for cell in ws[1]:
		cell.fill = green_fill
	# Auto-fit column widths
	for col in ws.columns:
		max_length = 0
		column = col[0].column_letter  # Get the column letter (e.g., "A", "B", etc.)
		for cell in col:
			try:
				# Calculate the length of the cell content
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		# Set the column width to fit the longest content, add padding
		adjusted_width = (max_length + 2) * 1.2  # Add some padding
		ws.column_dimensions[column].width = adjusted_width

	# Save the workbook to a BytesIO object
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file



@frappe.whitelist()
def get_item_details(args, doc=None, for_validate=False, overwrite_warehouse=True):
	"""
	args = {
			"item_code": "",
			"warehouse": None,
			"customer": "",
			"conversion_rate": 1.0,
			"selling_price_list": None,
			"price_list_currency": None,
			"plc_conversion_rate": 1.0,
			"doctype": "",
			"name": "",
			"supplier": None,
			"transaction_date": None,
			"conversion_rate": 1.0,
			"buying_price_list": None,
			"is_subcontracted": 0/1,
			"ignore_pricing_rule": 0/1
			"project": ""
			"set_warehouse": ""
	}
	"""
	from erpnext.stock.get_item_details import (process_args, process_string_args, get_basic_details, validate_item_details, get_item_tax_template, get_item_tax_map, get_party_item_code, set_valuation_rate, update_party_blanket_order, get_price_list_rate, get_pos_profile_item_details, update_bin_details, get_pricing_rule_for_item, update_stock, get_default_bom, get_gross_profit, remove_standard_fields)
	from frappe.utils import add_days, cint, flt
	args = process_args(args)
	for_validate = process_string_args(for_validate)
	overwrite_warehouse = process_string_args(overwrite_warehouse)
	item = frappe.get_cached_doc("Item", args.item_code)
	validate_item_details(args, item)

	if isinstance(doc, str):
		doc = json.loads(doc)

	if doc:
		args["transaction_date"] = doc.get("transaction_date") or doc.get("posting_date")

		if doc.get("doctype") == "Purchase Invoice":
			args["bill_date"] = doc.get("bill_date")

	out = get_basic_details(args, item, overwrite_warehouse)

	get_item_tax_template(args, item, out)
	out["item_tax_rate"] = get_item_tax_map(
		args.company,
		args.get("item_tax_template")
		if out.get("item_tax_template") is None
		else out.get("item_tax_template"),
		as_json=True,
	)

	get_party_item_code(args, item, out)

	if args.get("doctype") in ["Sales Order", "Quotation"]:
		set_valuation_rate(out, args)

	update_party_blanket_order(args, out)

	# Never try to find a customer price if customer is set in these Doctype
	current_customer = args.customer
	if args.get("doctype") in ["Purchase Order", "Purchase Receipt", "Purchase Invoice"]:
		args.customer = None

	out.update(get_price_list_rate(args, item))

	args.customer = current_customer

	if args.customer and cint(args.is_pos):
		out.update(get_pos_profile_item_details(args.company, args, update_data=True))

	if item.is_stock_item:
		update_bin_details(args, out, doc)

	# update args with out, if key or value not exists
	for key, value in out.items():
		if args.get(key) is None:
			args[key] = value

	data = get_pricing_rule_for_item(args, doc=doc, for_validate=for_validate)

	out.update(data)

	if (
		frappe.db.get_single_value("Stock Settings", "auto_create_serial_and_batch_bundle_for_outward")
		and not args.get("serial_and_batch_bundle")
		and (args.get("use_serial_batch_fields") or args.get("doctype") == "POS Invoice")
	):
		update_stock(args, out, doc)

	if args.transaction_date and item.lead_time_days:
		out.schedule_date = out.lead_time_date = add_days(args.transaction_date, item.lead_time_days)

	if args.get("is_subcontracted"):
		out.bom = args.get("bom") or get_default_bom(args.item_code)

	get_gross_profit(out)
	out.rate = args.rate or out.price_list_rate
	if args.doctype == "Material Request":
		out.amount = flt(args.qty) * flt(out.rate)

	out = remove_standard_fields(out)
	return out

# Payment Reconciliation
def reconcile_payments(company, account, customer):	
	outstanding_documents = get_outstanding_reference_documents({"party_type":'Customer', "party":customer, "party_account":account}) or 0
	flag = False
	if outstanding_documents:
		for i in outstanding_documents:
			if i.outstanding_amount > 0:
				flag = True
				break
	if flag:
		unallocated_amount = frappe.db.get_value("Payment Entry", [["party", "=", customer], ["unallocated_amount", ">", 0], ["docstatus", "=", 1]], "sum(unallocated_amount)") or 0
		credit_amount = frappe.db.get_value("Journal Entry Account", [["party", "=", customer], ["credit", ">", 0], ["reference_name", "=", None], ["docstatus", "=", 1]], "sum(credit)") or 0
		cn_amount = frappe.db.get_value("Sales Invoice", [["customer", "=", customer], ["outstanding_amount", "<", 0], ["is_return", "=", 1], ["docstatus", "=", 1]], "sum(outstanding_amount)") or 0
		if unallocated_amount or credit_amount or cn_amount:
			_cancel_old_reconciliations(company, account, customer)
			_create_reconciliation_entry(company, account, customer)

def _cancel_old_reconciliations(company, account, customer):
	reconciliations = frappe.get_all("Process Payment Reconciliation", filters={"party": customer, "company": company, "receivable_payable_account": account, "docstatus": 1, "status": ["!=", "Completed"]}, fields=["name"])
	for reconciliation in reconciliations:
		try:
			doc = frappe.get_doc("Process Payment Reconciliation", reconciliation.name)
			doc.cancel()
		except Exception as e:
			frappe.log_error("Error occurred while canceling reconciliation: {0}".format(str(e)))

def _create_reconciliation_entry(company, account, customer):
	reconciliation = frappe.get_doc({
		"doctype": "Process Payment Reconciliation",
		"party_type": "Customer",
		"party" : customer,
		"company": company,
		"receivable_payable_account": account,
		"default_advance_account": account
	})
	reconciliation.flags.ignore_permissions = True
	reconciliation.submit()

def update_remarks(doc_name, linked_doctype, affected_field, remarks):
	linked_doc = frappe.db.get_value(linked_doctype, {"custom_voucher_no": doc_name}, "name")
	if linked_doc:
		frappe.db.set_value("GL Entry", {"voucher_no": linked_doc}, "remarks", remarks)
		frappe.db.set_value(linked_doctype, linked_doc, affected_field, remarks)

# Roles Doctype
@frappe.whitelist()
def get_default_roles(role_type):
	return frappe.get_all("Libya Customizations Settings Roles", filters={"parentfield": role_type}, fields=["role"], pluck="role")
@frappe.whitelist()
def check_roles_included(role_type):
	roles = get_default_roles(role_type)
	if not roles:
		roles = get_default_roles_if_empty(role_type)
	user_roles = frappe.get_roles()
	return any(role in user_roles for role in roles)
@frappe.whitelist()
def get_default_roles_if_empty(role_type):
	roles = {
		"bulk_edit_prices": ["Chief Sales Officer"],
		"bypass_overdue_check": ["Chief Sales Officer"],
		"reserve_against_future_receipts": ["Chief Sales Officer"],
		"bypass_valuation_rate_check": ["Chief Sales Officer"],
		"bypass_price_list_check": ["Chief Sales Officer", "Sales Supervisor"],
		"show_valuation_rate": ["Chief Sales Officer"],
	}
	return roles.get(role_type, [])

def build_unreconcile_selection_map(selections, doctype, doc_name):
    selection_map = []

    if doctype in ["Sales Invoice", "Purchase Invoice"]:
        selection_map = [
            {
                "company": elem.get("company"),
                "voucher_type": elem.get("reference_doctype"),
                "voucher_no": elem.get("reference_name"),
                "against_voucher_type": doctype,
                "against_voucher_no": doc_name,
            }
            for elem in selections
        ]
        
    elif doctype in ["Payment Entry", "Journal Entry"]:
        selection_map = [
            {
                "company": elem.get("company"),
                "voucher_type": doctype,
                "voucher_no": doc_name,
                "against_voucher_type": elem.get("reference_doctype"),
                "against_voucher_no": elem.get("reference_name"),
            }
            for elem in selections
        ]

    return selection_map

def unreconcile_payments(voucher):
	for doc in frappe.get_all("Payment Entry", {"custom_voucher_no": voucher.name}, ["company", "name"]):
		linked_payments = get_linked_payments_for_doc(doc.company, "Payment Entry", doc.name)
		selection_map = build_unreconcile_selection_map(linked_payments, "Payment Entry", doc.name)
		create_unreconcile_doc_for_selection(json.dumps(selection_map))