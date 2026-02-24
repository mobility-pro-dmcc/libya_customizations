import frappe
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from frappe.utils.file_manager import save_file
from frappe.utils import get_site_path
import json
import math

@frappe.whitelist()
def increase_item_price(filters, percent):
    filters = json.loads(filters)
    items = frappe.get_all("Item Price", filters, ["stock_valuation_rate", "name"])
    for item in items:
        frappe.db.set_value("Item Price", item.name, "price_list_rate", math.ceil(item.stock_valuation_rate*(100+int(percent))/100))

@frappe.whitelist()
def export_item_price_data(filters):
    # Define the fields for "Item Price" export
    doctype = "Item Price"
    fields = ["name", "item_code", "item_name", "brand", "price_list_rate", "stock_valuation_rate", "stock_qty", "price_list"]
    filters = json.loads(filters)
    names = frappe.get_list(doctype, filters=filters)
    
    # Initialize the workbook and add header row
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Item Price Export"
    sheet.append(fields)  # Add headers based on fields

    # Define fill patterns
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")   # Light blue

    # Apply green fill to all header cells except "price_list_rate"
    price_list_rate_index = fields.index("price_list_rate") + 1  # Column index of "price_list_rate"
    for col_num, header in enumerate(fields, start=1):
        header_cell = sheet.cell(row=1, column=col_num)
        if col_num == price_list_rate_index:
            header_cell.fill = blue_fill  # Blue fill for "price_list_rate" header
        else:
            header_cell.fill = green_fill  # Green fill for other headers

    # Iterate over each selected document
    for name in names:
        # Fetch the document
        item_price_doc = frappe.get_doc(doctype, name)
        
        # Prepare a row with the specified fields
        row = [getattr(item_price_doc, field, "") for field in fields]
        sheet.append(row)  # Append the row to the sheet

        # Apply blue fill to the "price_list_rate" cell
        price_list_rate_cell = sheet.cell(row=sheet.max_row, column=price_list_rate_index)
        price_list_rate_cell.fill = blue_fill

    # Auto-fit column widths to fit the content
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter (e.g., "A", "B", etc.)
        for cell in col:
            try:
                # Calculate the length of the cell content
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Set the column width to fit the longest content, add padding
        sheet.column_dimensions[column].width = max_length + 2  # Add some padding

    # Define and save the Excel file path
    file_path = get_site_path("private", "files", "Item_Price_Export.xlsx")
    workbook.save(file_path)

    # Save file in Frappe's File Manager to generate download URL
    with open(file_path, "rb") as file:
        file_doc = save_file(
            "Item Price Export.xlsx",
            file.read(),
            "File",
            frappe.session.user,
            is_private=True
        )

    return file_doc.file_url


from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
@frappe.whitelist(allow_guest=True)
def import_item_price_data(file_url):
    
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    content = file_doc.get_content()
    data = read_xlsx_file_from_attached_file(fcontent=content)
    for row in data[1:]:
        frappe.db.set_value("Item Price", row[0], "price_list_rate", row[4])
    return data[1:]

@frappe.whitelist()
def update_stock_valuation_rate():
    item_prices = frappe.get_all("Item Price", fields=["name", "item_code", "production_year"])
    
    for ip in item_prices:
        # Get Stock Valuation Rate (based on item_code)
        avg_rate = frappe.db.sql("""
            SELECT SUM(stock_value) / SUM(actual_qty)
            FROM `tabBin`
            WHERE item_code = %s AND actual_qty > 0
        """, (ip.item_code,))[0][0]

        # Get Stock Qty (based on item_code + production_year)
        stock_qty = frappe.db.sql("""
            SELECT IFNULL(SUM(actual_qty), 0)
            FROM `tabStock Ledger Entry`
            WHERE is_cancelled = 0 AND item_code = %s AND IFNULL(production_year, '') = IFNULL(%s, '')
        """, (ip.item_code, ip.production_year))[0][0]

        qty_to_deliver = frappe.db.sql("""
            WITH reserved_qty AS (
                SELECT so.name, IF(SUM(soi.qty - soi.delivered_qty) > 0, SUM(soi.qty - soi.delivered_qty), 0) AS qty_to_deliver
                FROM `tabSales Order Item` soi
                INNER JOIN `tabSales Order` so ON soi.parent = so.name
                WHERE soi.docstatus = 1 AND so.docstatus = 1 AND so.status NOT IN ('Completed', 'Closed') AND soi.qty - soi.delivered_qty > 0
                AND soi.item_code = %s AND IFNULL(soi.production_year, '') = IFNULL(%s, '')
            )
            SELECT IFNULL(SUM(qty_to_deliver), 0)
            FROM reserved_qty
        """, (ip.item_code, ip.production_year))[0][0]

        available_qty = stock_qty - qty_to_deliver

        frappe.db.set_value("Item Price", ip.name, {
            "stock_valuation_rate": avg_rate or 0,
            "stock_qty": stock_qty or 0,
            "available_qty": available_qty or 0
        })

    frappe.db.commit()